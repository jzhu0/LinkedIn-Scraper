'''
Created on Jul 21, 2017


@author: Jason Zhu
'''
DATASET_NAME = "COPY - Linkedin & Operational Full List - 7.23.2017.xlsx"
DATASET_SHEETNAME = "Sheet1"
DOWNLOAD_DIRECTORY = "C:/Users/LinkedIn profiles"
LINKEDIN_EMAIL = "j888230@mvrht.net"
LINKEDIN_PASS = "j888230@"   # 10minutemail generated

import time
from bs4 import BeautifulSoup
import requests
import openpyxl
from selenium import webdriver
import ctypes

# Use ctypes to simulate keyboard presses - deal with the file Save As dialog box
SendInput = ctypes.windll.user32.SendInput

PUL = ctypes.POINTER(ctypes.c_ulong)
class KeyBdInput(ctypes.Structure):
    _fields_ = [("wVk", ctypes.c_ushort),
                ("wScan", ctypes.c_ushort),
                ("dwFlags", ctypes.c_ulong),
                ("time", ctypes.c_ulong),
                ("dwExtraInfo", PUL)]

class HardwareInput(ctypes.Structure):
    _fields_ = [("uMsg", ctypes.c_ulong),
                ("wParamL", ctypes.c_short),
                ("wParamH", ctypes.c_ushort)]

class MouseInput(ctypes.Structure):
    _fields_ = [("dx", ctypes.c_long),
                ("dy", ctypes.c_long),
                ("mouseData", ctypes.c_ulong),
                ("dwFlags", ctypes.c_ulong),
                ("time",ctypes.c_ulong),
                ("dwExtraInfo", PUL)]

class Input_I(ctypes.Union):
    _fields_ = [("ki", KeyBdInput),
                 ("mi", MouseInput),
                 ("hi", HardwareInput)]

class Input(ctypes.Structure):
    _fields_ = [("type", ctypes.c_ulong),
                ("ii", Input_I)]

def PressKey(hexKeyCode):
    extra = ctypes.c_ulong(0)
    ii_ = Input_I()
    ii_.ki = KeyBdInput( hexKeyCode, 0x48, 0, 0, ctypes.pointer(extra) )
    x = Input( ctypes.c_ulong(1), ii_ )
    ctypes.windll.user32.SendInput(1, ctypes.pointer(x), ctypes.sizeof(x))

# Extraction functions
def sign_in(driver):
    driver.find_element_by_id("login-email").send_keys(LINKEDIN_EMAIL)
    driver.find_element_by_id("login-password").send_keys(LINKEDIN_PASS)
    driver.find_element_by_id("login-submit").click()

def download_profile(driver):
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.5)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.5)
    driver.find_element_by_class_name("pv-top-card-section__overflow-wrapper").click()
    driver.find_element_by_class_name("save-to-pdf").click()

def set_download_name(pID):
    time.sleep(0.1)
    PressKey(0x24)   # HOME button
    # iterate over digits in pID, press each
    for c in map(int, str(pID)):
        code = 0x30 + c
        time.sleep(0.1)
        PressKey(code)
    time.sleep(0.1)
    PressKey(0xBD)   # - button
    time.sleep(0.1)
    PressKey(0x0D)   # ENTER button
    time.sleep(0.1)
    print "  ...Done with #" + str(pID)

print "Loading data set " + DATASET_NAME + "..."
BgOpData = openpyxl.load_workbook(DATASET_NAME)
links = {}
ids = {}
if (DATASET_SHEETNAME in BgOpData.sheetnames):
    BgOpDataSheet = BgOpData[DATASET_SHEETNAME]
else:
    raise Exception("Data set did not contain correct data sheet")
print "    Data set successfully opened"
for row in BgOpDataSheet.iter_rows():
    if (str(row[0].value) == "Project ID"):   # skip the first row
        continue
    
#    if (row[0].value < 0):   # Min row
#        continue
    if (row[0].value > 50):   # Max row
        break
    
    if not (str(row[2].value) == "None" or str(row[2].value) == "" or str(row[2].value) == "Private profile"):   # skip rows without a link
        links[str(row[1].value)] = str(row[2].value)
        ids[str(row[1].value)] = row[0].value
        print "Loaded pID #" + str(row[0].value) + ": " + str(row[1].value) + " " + str(row[2].value)
print "    Data set successfully loaded"

# test with smaller links dict:
#print "Using hard-coded data set."
#links = {"Raanan Zehavi" : "https://www.linkedin.com/in/raanan-zehavi-418933b/",
#"Aksel Chernitzky" : "https://www.linkedin.com/in/aksel-chernitzky-5902a/",
#"Janice Chernitzky" : "https://www.linkedin.com/in/janice-chernitzky-7259b763/",
#"Daniel Ferrazzoli" : "https://www.linkedin.com/in/daniel-ferrazzoli-2a76a72a/"}

# Set download location preferences
preferences = {"download.default_directory" : DOWNLOAD_DIRECTORY}
options = webdriver.ChromeOptions()
options.add_experimental_option("prefs", preferences)
driver = webdriver.Chrome(chrome_options = options)

# navigate to LinkedIn homepage - log in manually
driver.get("https://www.linkedin.com/")
sign_in(driver)
time.sleep(1.0)

for name, pURL in links.iteritems():
    pID = ids[name]
    print "Processing pID #" + str(pID) + ": " + name
    req = requests.get(pURL)
    data = BeautifulSoup(req.content, "html.parser")
    time.sleep(0.5)
    
    driver.get(pURL)
    download_profile(driver)
    time.sleep(2.0)
    set_download_name(pID)
    time.sleep(1.0)

print "    All data successfully parsed."
driver.quit()